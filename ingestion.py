"""
Background ingestion service.

Place REAL LabVIEW-generated .xlsx files in watched_folder/.
The service waits until a file is stable, extracts metadata, copies the
original workbook to report_storage/, and inserts searchable metadata
into SQLite.

The exact Excel field locations are intentionally NOT hard-coded yet.
Run inspect_excel.py on a real report first; then update the mapping
section below if the workbook uses non-adjacent/custom cells.
"""

import re
import shutil
import sqlite3
import time
from datetime import datetime, date
from pathlib import Path

from openpyxl import load_workbook

BASE_DIR = Path(__file__).resolve().parent
WATCH_DIR = BASE_DIR / "watched_folder"
STORAGE_DIR = BASE_DIR / "report_storage"
DB_PATH = BASE_DIR / "test_reports.db"

CUSTOMER_USERNAME = "tvs"

WATCH_DIR.mkdir(exist_ok=True)
STORAGE_DIR.mkdir(exist_ok=True)

SERIAL_LABELS = ["serial number", "serial no", "serial", "barcode", "qr code"]
RESULT_LABELS = ["overall result", "test result", "result", "status", "pass/fail", "pass fail"]
DATE_LABELS = ["test date", "date", "timestamp", "test time", "time"]
IQUBE_LABELS = ["i-qube", "iqube", "recipe", "recipe name"]


def normalize(value):
    return re.sub(r"\s+", " ", str(value or "").strip()).lower()


def find_value_near_label(ws, labels):
    for row in ws.iter_rows():
        for cell in row:
            label = normalize(cell.value)
            if not label:
                continue
            if any(target == label or target in label for target in labels):
                # Same row: first non-empty cell to the right.
                for col in range(cell.column + 1, min(ws.max_column, cell.column + 5) + 1):
                    value = ws.cell(cell.row, col).value
                    if value not in (None, ""):
                        return value
                # Same column: first non-empty cell below.
                for row_num in range(cell.row + 1, min(ws.max_row, cell.row + 5) + 1):
                    value = ws.cell(row_num, cell.column).value
                    if value not in (None, ""):
                        return value
    return None


def extract_metadata(path):
    wb = load_workbook(path, data_only=True, read_only=True)

    ws = wb.worksheets[0]

    # Actual LabVIEW report locations
    iqube = ws["M4"].value
    raw_date = ws["S3"].value
    serial = ws["M7"].value
    result = ws["S5"].value

    wb.close()

    if not iqube:
        raise ValueError("I-QUBE not found in cell M4")

    if not raw_date:
        raise ValueError("Date not found in cell S3")

    if not serial:
        raise ValueError("Serial/QR code not found in cell M7")

    if not result:
        raise ValueError("PASS/FAIL result not found in cell S5")

    # Convert date to database format YYYY-MM-DD
    if isinstance(raw_date, (datetime, date)):
        report_date = raw_date.strftime("%Y-%m-%d")
    else:
        parsed = None

        for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%m/%d/%Y", "%d-%m-%Y", "%m-%d-%Y"):
            try:
                parsed = datetime.strptime(str(raw_date).strip(), fmt)
                break
            except ValueError:
                pass

        if parsed:
            report_date = parsed.strftime("%Y-%m-%d")
        else:
            report_date = str(raw_date).strip()

    result = str(result).strip().upper()

    if "PASS" in result:
        result = "PASS"
    elif "FAIL" in result:
        result = "FAIL"

    return {
        "iqube": str(iqube).strip(),
        "report_date": report_date,
        "serial": str(serial).strip(),
        "result": result,
    }


def stable(path, wait_seconds=2):
    size1 = path.stat().st_size
    time.sleep(wait_seconds)
    return path.exists() and path.stat().st_size == size1


def customer_id():
    conn = sqlite3.connect(DB_PATH)
    row = conn.execute(
        "SELECT id FROM customers WHERE username=?", (CUSTOMER_USERNAME,)
    ).fetchone()
    conn.close()
    if not row:
        raise RuntimeError("Customer 'tvs' not found. Start app.py once to initialize the DB.")
    return row[0]


def ingest(path):
    metadata = extract_metadata(path)
    cid = customer_id()

    safe_result = re.sub(r"[^A-Za-z0-9_-]", "_", metadata["result"])
    safe_iqube = re.sub(r"[^A-Za-z0-9_.-]", "_", metadata["iqube"])
    destination_dir = STORAGE_DIR / CUSTOMER_USERNAME / safe_iqube / safe_result / metadata["report_date"]
    destination_dir.mkdir(parents=True, exist_ok=True)

    destination = destination_dir / path.name
    if destination.resolve() != path.resolve():
        shutil.copy2(path, destination)

    relative_storage_path = destination.relative_to(BASE_DIR)

    conn = sqlite3.connect(DB_PATH)
    conn.execute(
        """INSERT OR IGNORE INTO reports
           (customer_id, iqube, report_date, serial_number, result,
            original_filename, storage_path)
           VALUES (?, ?, ?, ?, ?, ?, ?)""",
        (
            cid,
            metadata["iqube"],
            metadata["report_date"],
            metadata["serial"],
            metadata["result"],
            path.name,
            str(relative_storage_path),
        )
    )
    conn.commit()
    conn.close()

    print(
        f"INGESTED | I-QUBE={metadata['iqube']} | "
        f"DATE={metadata['report_date']} | SERIAL={metadata['serial']} | "
        f"RESULT={metadata['result']}"
    )


def main():
    print(f"Watching: {WATCH_DIR}")
    processed = set()

    while True:
        for path in WATCH_DIR.glob("*.xlsx"):
            key = (str(path), path.stat().st_mtime_ns, path.stat().st_size)
            if key in processed:
                continue

            try:
                if stable(path):
                    ingest(path)
                    processed.add(key)
            except Exception as exc:
                print(f"ERROR | {path.name} | {exc}")

        time.sleep(2)


if __name__ == "__main__":
    main()
