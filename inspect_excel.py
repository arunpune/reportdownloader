"""Print workbook sheets and non-empty cells from a real Excel report."""

import sys
from pathlib import Path
from openpyxl import load_workbook

if len(sys.argv) != 2:
    print("Usage: python inspect_excel.py path/to/report.xlsx")
    raise SystemExit(1)

path = Path(sys.argv[1])
if not path.is_file():
    print(f"File not found: {path}")
    raise SystemExit(1)

wb = load_workbook(path, data_only=False, read_only=True)

for ws in wb.worksheets:
    print(f"\n===== SHEET: {ws.title} =====")
    for row in ws.iter_rows():
        values = []
        for cell in row:
            if cell.value not in (None, ""):
                values.append(f"{cell.coordinate}={cell.value!r}")
        if values:
            print(" | ".join(values))

wb.close()
